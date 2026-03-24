import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def print_report(results):
    print(f"\n{'='*100}")
    print(f"{'SCHEME':<15} {'TEST CASE':<30} {'PASS/FAIL':<10} {'ANALYSIS'}")
    print(f"{'='*100}")

    for r in results:
        print(f"{r['scheme']:<15} {r['test_case']:<30} {r['pass_fail']:<10} {r['analysis']}")

    total = len(results)
    passed = sum(1 for r in results if r["pass_fail"] == "PASS")
    failed = total - passed
    bugs = sum(1 for r in results if "BUG" in r["analysis"])

    print(f"\n{'='*100}")
    print(f"Total: {total} | Passed: {passed} | Failed: {failed} | Bugs: {bugs}")
    print(f"{'='*100}\n")


def export_excel_report(results, driver_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Auth Test Results"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "S.No",
        "Auth Scheme",
        "Test Case",
        "Connection String",
        "Pass/Fail",
        "Driver Response / Error Message",
        "Comments",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    for idx, r in enumerate(results, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=r["scheme"]).border = thin_border
        ws.cell(row=row, column=3, value=r["test_case"]).border = thin_border

        conn_cell = ws.cell(row=row, column=4, value=r["connection_string"])
        conn_cell.border = thin_border
        conn_cell.alignment = Alignment(wrap_text=True)

        pf_cell = ws.cell(row=row, column=5, value=r["pass_fail"])
        pf_cell.border = thin_border
        pf_cell.alignment = Alignment(horizontal="center")
        pf_cell.fill = pass_fill if r["pass_fail"] == "PASS" else fail_fill

        err_cell = ws.cell(row=row, column=6, value=r["error_message"])
        err_cell.border = thin_border
        err_cell.alignment = Alignment(wrap_text=True)

        ws.cell(row=row, column=7, value=r["comment"]).border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 40

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    total = len(results)
    passed = sum(1 for r in results if r["pass_fail"] == "PASS")
    failed = total - passed
    bugs = sum(1 for r in results if "BUG" in r["analysis"])

    summary_data = [
        ["Driver", driver_name.capitalize()],
        ["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Test Cases", total],
        ["Passed", passed],
        ["Failed", failed],
        ["Bugs Found", bugs],
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell_label = ws2.cell(row=row_idx, column=1, value=label)
        cell_label.font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=value)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 30

    # Save
    os.makedirs("reports", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reports/{driver_name}_{timestamp}.xlsx"
    wb.save(filename)
    print(f"[+] Excel report saved: {filename}")
